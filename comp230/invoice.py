from openpyxl import Workbook
wb = Workbook()
ws = wb.active

cust_ws = wb.create_sheet('Customers')
invoice_ws = wb.create_sheet('Invoice No.')
product_ws = wb.create_sheet('Products')
line_items_ws = wb.create_sheet('Line Items')

cust_ws['A1'] = 'Customer'
cust_ws['B1'] = 'Address'
cust_ws['C1'] = 'Phone Number'
cust_ws['D1'] = 'Contact Name'

invoice_ws['A1'] = 'Invoice No.'
invoice_ws['B1'] = 'Customer'
invoice_ws['C1'] = 'Date'
invoice_ws['D1'] = 'Amount'

product_ws['A1'] = 'Invoice No.'
product_ws['B1'] = 'Job description'
product_ws['C1'] = 'Cost per unit'

line_items_ws['A1'] = 'Invoice No.'
line_items_ws['B1'] = 'Product'
line_items_ws['C1'] = 'No. of Units'



cust_rows = [
    ('workforce', '55 st', '222-222-2222', 'joyce'),
    ('testcase', '66blvd', '222-111-1241', 'david'),
    ('nwt', 'something ave', '403-141-2144', 'guy'),
    ('1st', '111st', '226-111-2414', 'carrie'),
]

invoice_rows = [
    (1, 'Workforce', 'May 5, 2019', 1200),
    (2, 'testcase', 'April 23, 2019', 1500),
    (3, '1st', 'july 3, 2019', 50),
]

product_rows = [
    (1, 'Detail', 50),
    (2, 'Extra', 100),
]

line_items_rows = [
    (1, 'Wash', 24),
    (2, 'Wax', 15),
]

for row in cust_rows:
    cust_ws.append(row)

for row in invoice_rows:
    invoice_ws.append(row)

for row in product_rows:
    product_ws.append(row)

for row in line_items_rows:
    line_items_ws.append(row)



def final_invoice(inv):
    for row in invoice_ws.iter_rows():
        for x in row:
            if x.value == inv:
                print('Invoice No.: ',invoice_ws.cell(row=x.row, column=1).value)
                customer = invoice_ws.cell(row=x.row, column=2).value
    
    for row in cust_ws.iter_rows():
        for x in row:
            if x.value == str(customer).lower():
                print('Company: ',cust_ws.cell(row=x.row, column=1).value)
                print('Address: ',cust_ws.cell(row=x.row, column=2).value)
                print('Phone: ',cust_ws.cell(row=x.row, column=3).value)
                print('Contact: ',cust_ws.cell(row=x.row, column=4).value)


    for row in product_ws.iter_rows():
        for x in row:
            if x.value == inv:
                print('Service: ',product_ws.cell(row=x.row, column=2).value)
                print('Amount: ',product_ws.cell(row=x.row, column=3).value)
    
    for row in line_items_ws.iter_rows():
        for x in row:
            if x.value == inv:
                print('Line Item: ',line_items_ws.cell(row=x.row, column=2).value)

    # for row in cust_ws.iter_rows():
    #     for y in row: 
                
final_invoice(2)