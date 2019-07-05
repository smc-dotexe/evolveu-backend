from openpyxl import Workbook
wb = Workbook()

ws = wb.active #set to 0 by default

ws1 = wb.create_sheet("Mysheet") #insert at the end(default)
ws2 = wb.create_sheet("Mysheet", 0)#insert at the first position

ws.title = "New Title"

ws.sheet_properties.tabColor = "1072BA"
ws3 = wb["New Title"]
# print(wb.sheetnames)

# for sheet in wb:
#     print(sheet.title)

source = wb.active 
target = wb.copy_worksheet(source)

c = ws['A4']
ws['A4'] = 4

d = ws.cell(row=4, column=2, value=10)
#when worksheet is created in memory, it contains no cells. They are created
#when first accessed

cell_range = ws['A1':'C2']


