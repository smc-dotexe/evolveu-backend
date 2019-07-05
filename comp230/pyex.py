import openpyxl

book = openpyxl.load_workbook('invoice_spreadsheet2.xlsx')

active_sheet = book.active
sheets = book.sheetnames
rows = active_sheet.max_row
columns = active_sheet.max_column
a_none = None




# customer_list = []
# invoice_list = []
# product_list = []
# line_item_list = []


#make a clas
#user instantiates the class by invoice number
#the class builds the spreadsheet
#grabs the invoice number which has the items, price and total 
#prints the items to a file 
class Customer_class:
    def __init__(self):
        self.customer_list = []
        customer_sheet = book[sheets[0]]
        for r in range(1, customer_sheet.max_row + 1):
            customer_dict={}
            for c in range(1, customer_sheet.max_column + 1):
                titles = customer_sheet.cell(row=1, column=c).value
                items = customer_sheet.cell(row=r, column=c).value
                if titles != None and items != None:
                    customer_dict[titles] = items
                else: return
            self.customer_list.append(customer_dict)

        
class Invoice_class:
    def __init__(self):
        self.invoice_list = []
        invoice_sheet = book[sheets[1]]
        for r in range(1, invoice_sheet.max_row + 1):
            invoice_dict={}
            for c in range (1, invoice_sheet.max_column + 1):
                titles = invoice_sheet.cell(row=1, column=c).value
                items = invoice_sheet.cell(row=r, column=c).value
                if titles != None and items != None: 
                    invoice_dict[titles] = items
                else: return
            self.invoice_list.append(invoice_dict)


class Line_item_class:
    def __init__ (self):
        self.line_item_list = []
        line_item_sheet = book[sheets[2]]
        for r in range(1, line_item_sheet.max_row + 1):
            line_item_dict={}
            for c in range(1, line_item_sheet.max_column + 1):
                titles = line_item_sheet.cell(row=1, column=c).value
                items = line_item_sheet.cell(row=r, column=c).value
                line_item_dict[titles] = items
            self.line_item_list.append(line_item_dict)



class Product_class:
    def __init__ (self):
        self.product_list = []
        product_sheet = book[sheets[3]]
        for r in range(2, product_sheet.max_row + 1):
            product_dict = {}
            for c in range(2, product_sheet.max_column + 1):
                titles = product_sheet.cell(row=1, column=c).value
                items = product_sheet.cell(row=r, column=c).value
                product_dict[titles] = items
            self.product_list.append(product_dict)
    
customer = Customer_class()
invoice = Invoice_class()
line_item = Line_item_class()
product = Product_class()

print(product.product_list)
line_product_list = []
product_total_list = []
line_units_list = []
product_price_list=[]

#function to tell if the input number is in the invoice no. list
def check_num(a):
    check_invoice_num = []
    for x, i in enumerate(invoice.invoice_list):
        invoice_values = invoice.invoice_list[x].values()
        check_invoice_list = list(invoice_values)
        check_invoice_num.append(check_invoice_list[0])
    
    if a in check_invoice_num:
        return True
    else:
        return False

# def check_blanks(param): 
#     master_list = []  
#     for x in param:
#         for i, j in x.items():
#             master_list.append(i)
#             master_list.append(j)
#     print(master_list)
    

# check_blanks()
        

def create_invoice(num):
    print('num ', num)
    a = None
    if check_num(num) == True: 
        #invoice sheet
        for x, i in enumerate(invoice.invoice_list):
            invoice_list_values = invoice.invoice_list[x].values()
            invoice_list_number = list(invoice_list_values)
            invoice_num = invoice_list_number[0]
            #check if there are any blank inputs in Invoice No. and Customer No.
            # if a == invoice_num or a == invoice_list_number[1]:
            #     return print('blank in invoice sheet')
            if num == invoice_num:
                matched_invoice = invoice_num
                invoice_cust_num = invoice_list_number[1]
                invoice_date = invoice_list_number[2]
     
        #customer sheet    
        for cust_index, cust_value in enumerate(customer.customer_list):   
            customer_list_values = customer.customer_list[cust_index].values()
            customer_list_number = list(customer_list_values)
            customer_num = customer_list_number[0]

           
            if invoice_cust_num == customer_num:
                matched_cust_num = customer_num
                customer_name = customer_list_number[1]
                customer_address = customer_list_number[2]
        #line item sheet 
        for line_index, line_value in enumerate(line_item.line_item_list):
            line_list_values = line_item.line_item_list[line_index].values()
            line_list_number = list(line_list_values)

            if matched_invoice == line_list_number[0]:
                line_product_number = line_list_number[1]
                line_units = line_list_number[2]
                line_product_list.append(line_product_number)
                line_units_list.append(line_units)
                for product_index, product_value in enumerate(product.product_list):
                    product_values = product.product_list[product_index].values()
                    product_list_number = list(product_values)
                    if line_product_number == product_list_number[0]:
                        product_name = product_list_number[1]
                        product_price = product_list_number[2]
                        product_total_list.append(product_name)
                        product_price_list.append(product_price)

        with open('invoice.txt', 'w') as f:
            f.write(f'Invoice No.: {matched_invoice}\n'\
                    f'Date: {invoice_date}\n' \
                    f'{customer_name}\n{customer_address}\n')
            f.write('-----------------------------------\n')
            f.write(f'Product: {", ".join([str(x) for x in product_total_list])}\n' \
                    f'Price: {", ".join([str(x)for x in product_price_list])}\n' \
                    f'Quantity: {", ".join([str(x) for x in line_units_list])}\n')
            f.write('-----------------------------------\n')
            f.write(f'Total: ${grand_total()}')
            f.close() 
    else: 
        return print("can't find invoice")

def invoice_totals():
    return [j for i in zip(product_price_list,line_units_list) for j in i]
    
def grand_total():
    total = 0
    for a, b in zip(product_price_list, line_units_list):
        c = a*b
        total += c
    return round(total, 2)

create_invoice(34)



